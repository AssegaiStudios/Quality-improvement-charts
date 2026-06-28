"""Excel workbook I/O for pyqicharts Excel Companion."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from .config import CONFIG_FIELDS, config_defaults_frame, parse_config_frame


SHEETS = [
    "Data",
    "Config",
    "Chart",
    "ChartData",
    "SPCSummary",
    "Signals",
    "NHSInterpretation",
    "PowerBI",
    "Exports",
    "Log",
    "Help",
]


def _require_openpyxl():
    """Import openpyxl lazily so normal package imports stay light."""

    try:
        import openpyxl
        from openpyxl import Workbook, load_workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError as exc:
        raise ImportError("openpyxl is required for workbook file operations. Install pyqicharts[excel].") from exc
    return openpyxl, Workbook, load_workbook, ExcelImage, Font, PatternFill, dataframe_to_rows


def sample_excel_data() -> pd.DataFrame:
    """Return sample rows covering common Excel companion chart options."""

    return pd.DataFrame(
        {
            "period": [f"Month {i}" for i in range(1, 13)],
            "value": [10, 11, 9, 10, 12, 11, 10, 9, 13, 14, 13, 15],
            "denominator": [100] * 12,
            "expected": [11] * 12,
            "subgroup": [1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4, 4],
            "phase": ["Baseline"] * 6 + ["Improvement"] * 6,
            "exclude": [""] * 12,
            "target": [14] * 12,
            "intervention": [""] * 6 + ["New process"] + [""] * 5,
            "annotation": [""] * 12,
            "category": ["Delay", "Delay", "Error", "Delay", "Access", "Error", "Delay", "Access", "Other", "Delay", "Error", "Access"],
            "count": [4, 3, 2, 5, 3, 2, 4, 3, 1, 6, 2, 4],
        }
    )


def _append_frame(ws: Any, frame: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet from A1."""

    _, _, _, _, _, _, dataframe_to_rows = _require_openpyxl()
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)


def create_template(path: str | Path, macro_enabled: bool | None = None) -> Path:
    """Create a pyqicharts Excel Companion workbook template."""

    _, Workbook, _, _, Font, PatternFill, _ = _require_openpyxl()
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for name in SHEETS:
        workbook.create_sheet(name)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet["A1"].font = Font(bold=True)

    data = workbook["Data"]
    _append_frame(data, sample_excel_data())
    for cell in data[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    config = workbook["Config"]
    defaults = config_defaults_frame()
    default_values = {
        "data_sheet": "Data",
        "data_range": "",
        "x_column": "period",
        "y_column": "value",
        "chart_type": "run",
        "denominator_column": "denominator",
        "expected_column": "expected",
        "subgroup_column": "subgroup",
        "target_value": "",
        "target_column": "target",
        "recalculate_after": "",
        "theme": "default",
        "rules": "nhs",
        "method": "anhoej",
        "direction": "high is good",
        "export_png": "FALSE",
        "export_excel": "FALSE",
        "export_powerpoint": "FALSE",
        "export_bundle": "FALSE",
        "debug_mode": "FALSE",
        "powerbi_enabled": "TRUE",
    }
    defaults["value"] = defaults.apply(lambda row: default_values.get(row["field"], row["value"]), axis=1)
    _append_frame(config, defaults)
    for cell in config[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    config.column_dimensions["A"].width = 28
    config.column_dimensions["B"].width = 34

    chart = workbook["Chart"]
    chart["A1"] = "pyqicharts Excel Companion"
    chart["A2"] = "Use the xlwings macros listed below, or run pyqicharts-excel-init to create a fresh workbook."
    for row, label in enumerate(
        ["Generate Chart", "Generate All Outputs", "Export Report Bundle", "Clear Outputs", "Validate Workbook"],
        start=4,
    ):
        chart[f"A{row}"] = label
        chart[f"B{row}"] = f"Assign this button to pyqicharts_excel.runner.{label.lower().replace(' ', '_')}()"

    help_sheet = workbook["Help"]
    help_lines = [
        "How to use this workbook",
        "1. Paste tabular data into the Data sheet. Keep headers in row 1.",
        "2. Edit the Config sheet. For most SPC charts set x_column, y_column and chart_type.",
        "3. For P and U charts set denominator_column. For P' and U' charts set expected_column.",
        "4. Run Generate Chart through xlwings. Outputs are written to Chart, ChartData, SPCSummary, Signals, NHSInterpretation, PowerBI, Exports and Log.",
        "5. Install with: pip install pyqicharts pyqicharts-excel xlwings",
        "6. Development install: pip install -e \".[excel,dev]\" then xlwings addin install.",
        "Troubleshooting: check Log first. Set debug_mode TRUE to include technical traceback details.",
        "Ribbon/add-in Phase 2 path: see docs/excel_setup.md and pyqicharts_excel.ribbon.",
    ]
    for row, line in enumerate(help_lines, start=1):
        help_sheet.cell(row=row, column=1, value=line)
    help_sheet.column_dimensions["A"].width = 120

    workbook.save(output)
    return output


def create_templates(folder: str | Path = "excel") -> dict[str, Path]:
    """Create xlsx and xlsm workbook templates plus the Excel README."""

    target = Path(folder)
    target.mkdir(parents=True, exist_ok=True)
    xlsx = create_template(target / "pyqicharts_excel_template.xlsx")
    # The xlsm file is an xlwings-ready workbook shell. Users can add signed VBA
    # buttons in Excel; openpyxl cannot safely author VBA projects from scratch.
    xlsm = create_template(target / "pyqicharts_excel_template.xlsm", macro_enabled=True)
    readme = target / "README_excel.md"
    readme.write_text(
        "# pyqicharts Excel Companion\n\n"
        "Open the template, paste data into `Data`, edit `Config`, and run the xlwings entry points in `pyqicharts_excel.runner`.\n\n"
        "Install: `pip install pyqicharts pyqicharts-excel xlwings` and then `xlwings addin install`.\n",
        encoding="utf-8",
    )
    return {"xlsx": xlsx, "xlsm": xlsm, "readme": readme}


def read_config_from_workbook(path: str | Path):
    """Read the Config sheet from a workbook path."""

    frame = read_sheet_frame(path, "Config")
    return parse_config_frame(frame)


def _used_range(ws: Any) -> str:
    """Return the used worksheet range in A1 notation."""

    return f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"


def read_sheet_frame(path: str | Path, sheet_name: str, data_range: str = "") -> pd.DataFrame:
    """Read a rectangular Excel range into a DataFrame using row 1 as headers."""

    _, _, load_workbook, _, _, _, _ = _require_openpyxl()
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain sheet '{sheet_name}'.")
    ws = workbook[sheet_name]
    range_ref = data_range or _used_range(ws)
    cells = ws[range_ref]
    rows = [[cell.value for cell in row] for row in cells]
    rows = [row for row in rows if any(value is not None and value != "" for value in row)]
    if not rows:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return pd.DataFrame(rows[1:], columns=headers)


def write_frame_to_sheet(path: str | Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Replace a worksheet with a DataFrame."""

    _, _, load_workbook, _, Font, PatternFill, dataframe_to_rows = _require_openpyxl()
    workbook = load_workbook(path)
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = workbook.create_sheet(sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)
    if ws.max_row:
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
    workbook.save(path)


def write_log(path: str | Path, message: str, level: str = "INFO") -> None:
    """Append a timestamped log message to the Log sheet."""

    _, _, load_workbook, _, Font, _, _ = _require_openpyxl()
    workbook = load_workbook(path)
    ws = workbook["Log"] if "Log" in workbook.sheetnames else workbook.create_sheet("Log")
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(["timestamp", "level", "message"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
    ws.append([datetime.now().isoformat(timespec="seconds"), level, message])
    workbook.save(path)


def insert_chart_image(path: str | Path, image_path: str | Path, sheet_name: str = "Chart", title: str = "") -> None:
    """Insert a chart PNG into the configured Chart sheet."""

    _, _, load_workbook, ExcelImage, Font, _, _ = _require_openpyxl()
    workbook = load_workbook(path)
    ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = title or "pyqicharts chart"
    ws["A2"] = f"Generated {datetime.now().isoformat(timespec='seconds')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.add_image(ExcelImage(str(image_path)), "A4")
    workbook.save(path)


def clear_output_sheets(path: str | Path) -> None:
    """Clear generated output sheets while leaving Data, Config and Help intact."""

    _, _, load_workbook, _, _, _, _ = _require_openpyxl()
    workbook = load_workbook(path)
    for sheet_name in ["Chart", "ChartData", "SPCSummary", "Signals", "NHSInterpretation", "PowerBI", "Exports", "Log"]:
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        ws.delete_rows(1, ws.max_row)
    workbook.save(path)

