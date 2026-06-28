import importlib.util

import pytest

from pyqicharts_excel.excel_io import SHEETS, create_template, read_config_from_workbook, read_sheet_frame, write_frame_to_sheet


pytestmark = pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None, reason="openpyxl is not installed")


def test_template_contains_required_sheets(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    workbook = load_workbook(path)
    assert set(SHEETS).issubset(workbook.sheetnames)


def test_read_config_and_data_from_template(tmp_path):
    path = create_template(tmp_path / "template.xlsx")
    cfg = read_config_from_workbook(path)
    data = read_sheet_frame(path, cfg.data_sheet, cfg.data_range)
    assert cfg.chart_type == "run"
    assert {"period", "value", "denominator", "expected"}.issubset(data.columns)


def test_write_frame_to_sheet_replaces_content(tmp_path):
    import pandas as pd
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    write_frame_to_sheet(path, "ChartData", pd.DataFrame({"a": [1], "b": [2]}))
    workbook = load_workbook(path, data_only=True)
    assert workbook["ChartData"]["A1"].value == "a"
    assert workbook["ChartData"]["B2"].value == 2

