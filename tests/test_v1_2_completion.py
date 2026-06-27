import math

import pandas as pd
from openpyxl import load_workbook

from pyqicharts import (
    anhoej_rules,
    export_excel,
    intervention_metadata_table,
    kpi_table,
    nelson_rule_signals,
    nhs_interpretation_table,
    phase_metadata_table,
    powerbi_table,
    qic,
    qic_table,
    signal_table,
    spc_summary_table,
    target_metadata_table,
)


def test_anhoej_rules_use_exact_crossing_tail_and_ignore_ties_missing_exclusions():
    values = [1, 1, 1, 2, 2, None, 3, 3, 3]
    result = anhoej_rules(values, exclude=[False, True, False, False, False, False, False, False, False])
    assert result.median == 2
    assert result.n_used == 5
    assert result.crossings == 1
    assert result.lower_crossings_limit == 0
    assert not result.signal_few_crossings


def test_run_chart_methods_are_available_and_documented_in_table():
    df = pd.DataFrame({"month": range(1, 10), "value": [1, 1, 1, 1, 2, 9, 9, 9, 9]})
    for method in ["anhoej", "bestbox", "cutbox"]:
        out = qic_table(df, "month", "value", chart="run", method=method)
        assert out["run_chart_method"].iloc[0] == method
        assert "centre" in out


def test_segment_aware_nelson_uses_recalculated_centres_and_sigma():
    df = pd.DataFrame({"month": range(1, 21), "value": [10, 11] * 5 + [100, 101] * 5})
    out = qic_table(df, "month", "value", chart="i", recalculation_points=[11], rules="nelson")
    assert out["segment_id"].nunique() == 2
    assert out.groupby("segment_id")["centre"].first().nunique() == 2
    assert "signal_count" in out


def test_p_prime_and_u_prime_are_methodologically_distinct_with_denominator():
    df = pd.DataFrame({"month": [1, 2, 3, 4], "observed": [10, 12, 8, 11], "expected": [9, 11, 10, 10], "n": [100, 110, 95, 105]})
    p_out = qic_table(df, "month", "observed", chart="p_prime", expected="expected", denominator="n")
    u_out = qic_table(df, "month", "observed", chart="u_prime", expected="expected", denominator="n")
    assert p_out["centre_label"].iloc[0] == "Risk-adjusted proportion"
    assert u_out["centre_label"].iloc[0] == "Risk-adjusted rate"
    assert not p_out["ucl"].equals(u_out["ucl"])


def test_xbar_and_s_support_process_metadata_and_variable_subgroup_limits():
    df = pd.DataFrame({"batch": [1, 1, 2, 2, 2, 3, 3, 4, 4, 4], "value": [10, 11, 9, 10, 11, 20, 21, 19, 20, 21]})
    out = qic_table(df, "batch", "value", chart="xbar", break_points=[3], exclude=[2], phases=[{"start": 1, "label": "A"}, {"start": 3, "label": "B"}])
    assert out["segment_id"].nunique() == 2
    assert out.loc[out["batch"] == 2, "excluded"].iloc[0]
    assert out["ucl"].nunique() > 1
    s_out = qic_table(df, "batch", "value", chart="s", break_points=[3])
    assert (s_out["lcl"] >= 0).all()


def test_powerbi_tables_have_required_schema_columns():
    chart = qic(
        pd.DataFrame({"month": range(1, 14), "value": [10] * 12 + [20]}),
        "month",
        "value",
        chart="i",
        target=15,
        interventions=[{"point": 5, "label": "Change"}],
        phases=[{"start": 1, "label": "Baseline"}],
    )
    tables = [
        powerbi_table(chart),
        spc_summary_table(chart),
        signal_table(chart),
        kpi_table(chart),
        nhs_interpretation_table(chart),
        phase_metadata_table(chart),
        intervention_metadata_table(chart),
        target_metadata_table(chart),
    ]
    for table in tables:
        assert {"schema_version", "chart_id", "chart_type"}.issubset(table.columns)


def test_excel_export_contains_schema_and_metadata_sheets(tmp_path):
    chart = qic(
        pd.DataFrame({"month": range(1, 14), "value": [10] * 12 + [20]}),
        "month",
        "value",
        chart="i",
        target=15,
        interventions=[{"point": 5, "label": "Long intervention label for export validation"}],
        step_changes=[{"point": 8, "label": "Step"}],
        phases=[{"start": 1, "label": "Baseline"}],
    )
    path = export_excel(chart, tmp_path / "report.xlsx")
    workbook = load_workbook(path)
    assert {"Signal table", "KPI summary", "NHS interpretation", "Phase metadata", "Interventions", "Targets"}.issubset(workbook.sheetnames)


def test_rare_event_edge_cases_and_invalid_values():
    g = qic_table(pd.DataFrame({"x": [1], "y": [0]}), "x", "y", chart="g")
    assert math.isnan(g["ucl"].iloc[0]) or g["ucl"].iloc[0] >= 0
    t = qic_table(pd.DataFrame({"x": [1, 2, 3], "y": [0, 10, 200]}), "x", "y", chart="t")
    assert "signal" in t
    try:
        qic_table(pd.DataFrame({"x": [1], "y": [-1]}), "x", "y", chart="g")
    except ValueError as exc:
        assert "non-negative" in str(exc)
    else:
        raise AssertionError("negative G chart intervals should fail")


def test_nelson_missing_zero_sigma_and_flat_series_do_not_crash():
    assert nelson_rule_signals([1, None, 2], centre=1, sigma=0).empty
    assert nelson_rule_signals([1, 1, 1], centre=1, sigma=1).empty
