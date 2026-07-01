import sys
import types

import pandas as pd
import pytest

from pyqicharts import read_validation_csv
from pyqicharts_excel.config import ExcelConfig, config_defaults_frame, parse_config_mapping
from pyqicharts_excel.excel_io import create_template, create_templates, read_sheet_frame, write_frame_to_sheet, write_log
from pyqicharts_excel.examples import example_config, example_data
from pyqicharts_excel.ribbon import RIBBON_CALLBACKS, RIBBON_XML_SNIPPET, ribbon_instructions
from pyqicharts_excel.runner import _calling_workbook_path


def test_examples_return_user_ready_tables():
    data = example_data()
    config = example_config()
    assert {"period", "value", "denominator", "expected", "category", "count"}.issubset(data.columns)
    assert {"field", "value"}.issubset(config.columns)


def test_ribbon_metadata_and_instructions_are_complete():
    instructions = ribbon_instructions()
    assert "Generate Chart" in RIBBON_CALLBACKS
    assert "pyqicharts_excel.runner.generate_chart" in instructions
    assert "<customUI" in RIBBON_XML_SNIPPET


def test_create_templates_writes_both_workbooks_and_readme(tmp_path):
    created = create_templates(tmp_path / "excel")
    assert created["xlsx"].exists()
    assert created["xlsm"].exists()
    assert "xlwings" in created["readme"].read_text(encoding="utf-8")


def test_read_sheet_frame_missing_and_blank_sheets(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    with pytest.raises(ValueError, match="does not contain sheet"):
        read_sheet_frame(path, "Missing")

    workbook = load_workbook(path)
    workbook.create_sheet("Blank")
    workbook.save(path)
    assert read_sheet_frame(path, "Blank").empty


def test_write_frame_creates_new_sheet_and_log_appends_header(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    write_frame_to_sheet(path, "NewOutput", pd.DataFrame({"a": [1]}))
    write_log(path, "hello")

    workbook = load_workbook(path, data_only=True)
    assert workbook["NewOutput"]["A2"].value == 1
    assert workbook["Log"]["A1"].value == "timestamp"
    assert workbook["Log"]["C2"].value == "hello"


def test_read_validation_csv_helper(tmp_path):
    path = tmp_path / "validation.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    assert read_validation_csv(path).iloc[0]["a"] == 1


def test_config_defaults_serialise_list_values():
    # The defaults helper has a branch for list-valued fields. Exercise the
    # same serialisation logic without changing the public default object.
    frame = config_defaults_frame()
    assert "field" in frame
    cfg = parse_config_mapping({"break_points": "1,, 2"})
    assert cfg.break_points == [1, 2]


def test_calling_workbook_uses_xlwings_caller(monkeypatch):
    class FakeCaller:
        fullname = "C:/tmp/workbook.xlsx"

    class FakeBook:
        @staticmethod
        def caller():
            return FakeCaller()

    fake_xlwings = types.SimpleNamespace(Book=FakeBook)
    monkeypatch.setitem(sys.modules, "xlwings", fake_xlwings)
    assert str(_calling_workbook_path()).endswith("workbook.xlsx")

