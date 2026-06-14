import importlib.util
import json

import matplotlib
matplotlib.use("Agg")
import pandas as pd
import pytest

from pyqicharts import (
    create_report_bundle,
    export_excel,
    export_png,
    export_powerpoint,
    powerbi_table,
    qic,
    special_cause_summary_table,
    spc_summary_table,
)


def sample_chart():
    df = pd.DataFrame({"month": range(1, 14), "value": [10] * 12 + [20]})
    return qic(df, "month", "value", chart="i", improvement="high is good", target=15)


def test_png_file_is_created(tmp_path):
    chart = sample_chart()
    output = export_png(chart, tmp_path / "chart.png")
    assert output.exists()
    assert output.stat().st_size > 0


def test_chart_save_png_file_is_created(tmp_path):
    chart = sample_chart()
    output = chart.save_png(tmp_path / "chart-method.png")
    assert output.exists()
    assert output.stat().st_size > 0


@pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None, reason="openpyxl is not installed")
def test_excel_workbook_is_created_and_opens(tmp_path):
    from openpyxl import load_workbook

    chart = sample_chart()
    output = export_excel(chart, tmp_path / "report.xlsx")
    workbook = load_workbook(output)
    assert {"Chart data", "SPC summary", "Special causes"}.issubset(workbook.sheetnames)


@pytest.mark.skipif(importlib.util.find_spec("pptx") is None, reason="python-pptx is not installed")
def test_powerpoint_file_is_created_and_opens(tmp_path):
    from pptx import Presentation

    chart = sample_chart()
    output = export_powerpoint(chart, tmp_path / "report.pptx")
    presentation = Presentation(str(output))
    assert len(presentation.slides) >= 2


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None or importlib.util.find_spec("pptx") is None,
    reason="reporting extras are not installed",
)
def test_report_bundle_folder_contains_expected_files(tmp_path):
    chart = sample_chart()
    metadata = create_report_bundle([chart], tmp_path / "bundle")
    folder = tmp_path / "bundle"
    assert (folder / "chart_1.png").exists()
    assert (folder / "report.xlsx").exists()
    assert (folder / "report.pptx").exists()
    assert (folder / "metadata.json").exists()
    saved = json.loads((folder / "metadata.json").read_text(encoding="utf-8"))
    assert saved["chart_count"] == metadata["chart_count"] == 1


def test_powerbi_tables_return_dataframes():
    chart = sample_chart()
    assert isinstance(powerbi_table(chart), pd.DataFrame)
    assert isinstance(spc_summary_table(chart), pd.DataFrame)
    assert isinstance(special_cause_summary_table(chart), pd.DataFrame)
    assert spc_summary_table(chart)["signals"].iloc[0] > 0


def test_optional_dependencies_fail_gracefully_if_missing(monkeypatch, tmp_path):
    import builtins

    chart = sample_chart()
    original_import = builtins.__import__

    def blocked_import(name, *args, **kwargs):
        if name.startswith("openpyxl"):
            raise ImportError("blocked for test")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", blocked_import)
    with pytest.raises(ImportError, match="pip install pyqicharts\\[reporting\\]"):
        export_excel(chart, tmp_path / "report.xlsx")
