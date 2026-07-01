import importlib.util

import pytest

from pyqicharts_excel.excel_io import create_template
from pyqicharts_excel.runner import (
    _call_pyqicharts,
    _calling_workbook_path,
    _exclude_points,
    _interventions_from_columns,
    export_report_bundle,
    clear_outputs,
    generate_all_outputs,
    generate_chart,
    init_template,
    init_template_cli,
    main,
    refresh_config_defaults,
    validate_workbook,
)
from pyqicharts_excel.config import ExcelConfig


pytestmark = pytest.mark.skipif(importlib.util.find_spec("openpyxl") is None, reason="openpyxl is not installed")


def test_generate_chart_writes_outputs(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    message = generate_chart(path)
    assert "completed successfully" in message
    workbook = load_workbook(path, data_only=True)
    assert workbook["ChartData"]["A1"].value is not None
    assert workbook["SPCSummary"]["A1"].value == "schema_version"
    assert workbook["NHSInterpretation"]["A1"].value == "chart_id"


def test_validate_workbook_returns_success(tmp_path):
    path = create_template(tmp_path / "template.xlsx")
    assert validate_workbook(path) == "Workbook validation passed."


def test_clear_outputs_clears_generated_sheets(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    generate_chart(path)
    assert clear_outputs(path) == "Outputs cleared."
    workbook = load_workbook(path, data_only=True)
    assert workbook["ChartData"]["A1"].value is None


def test_runner_logs_friendly_error(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    workbook = load_workbook(path)
    ws = workbook["Config"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "chart_type":
            row[1].value = "p"
        if row[0].value == "denominator_column":
            row[1].value = ""
    workbook.save(path)

    message = generate_chart(path)
    assert "failed" in message
    workbook = load_workbook(path, data_only=True)
    log_values = [cell.value for cell in workbook["Log"]["C"] if cell.value]
    assert any("denominator" in value for value in log_values)


def test_generate_aliases_and_bundle_paths(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    assert generate_all_outputs(path) == "Generate chart completed successfully."
    assert main(path) == "Generate chart completed successfully."
    assert export_report_bundle(path) == "Export report bundle completed successfully."
    workbook = load_workbook(path, data_only=True)
    statuses = [cell.value for cell in workbook["Exports"]["A"] if cell.value]
    assert "export_type" in statuses


def test_refresh_config_defaults_and_init_template(tmp_path, capsys):
    path = create_template(tmp_path / "template.xlsx")
    assert refresh_config_defaults(path) == "Config defaults refreshed."

    created = init_template(tmp_path / "created.xlsx")
    assert created.exists()

    cwd_file = tmp_path / "cli_template.xlsx"
    old = __import__("os").getcwd()
    try:
        __import__("os").chdir(tmp_path)
        init_template_cli()
        captured = capsys.readouterr()
        assert "pyqicharts_excel_template.xlsx" in captured.out
        assert (tmp_path / "pyqicharts_excel_template.xlsx").exists()
    finally:
        __import__("os").chdir(old)


def test_validate_workbook_reports_errors(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    workbook = load_workbook(path)
    ws = workbook["Config"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "chart_type":
            row[1].value = "not_a_chart"
    workbook.save(path)

    message = validate_workbook(path)
    assert message.startswith("Validation failed")
    assert "not_a_chart" in message


def test_runner_debug_mode_logs_traceback(tmp_path):
    from openpyxl import load_workbook

    path = create_template(tmp_path / "template.xlsx")
    workbook = load_workbook(path)
    ws = workbook["Config"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "chart_type":
            row[1].value = "p"
        if row[0].value == "denominator_column":
            row[1].value = ""
        if row[0].value == "debug_mode":
            row[1].value = "TRUE"
    workbook.save(path)

    assert "failed" in generate_chart(path)
    workbook = load_workbook(path, data_only=True)
    log_text = "\n".join(str(cell.value) for cell in workbook["Log"]["C"] if cell.value)
    assert "Traceback" in log_text


def test_runner_private_helpers_cover_excel_config_branches():
    data = __import__("pandas").DataFrame(
        {
            "x": [1, 2, 3],
            "y": [1, 2, 3],
            "intervention": ["", "New pathway", ""],
            "exclude": ["", "yes", "0"],
            "count": [3, 2, 1],
        }
    )
    cfg = ExcelConfig(x_column="x", y_column="y", intervention_column="intervention", exclude_column="exclude")
    assert _interventions_from_columns(data, cfg) == [{"point": 2, "label": "New pathway"}]
    assert _exclude_points(data, cfg) == [2]
    assert _interventions_from_columns(data, ExcelConfig()) == []
    assert _exclude_points(data, ExcelConfig()) == []

    pareto = _call_pyqicharts(data, ExcelConfig(chart_type="pareto", x_column="x", denominator_column="count"))
    assert pareto.category == "x"

    chart = _call_pyqicharts(data, ExcelConfig(x_column="x", y_column="y", chart_type="xmr", freeze_after=2, break_points=[3]))
    assert chart.chart == "i"


def test_calling_workbook_without_path_raises_friendly_error():
    with pytest.raises(RuntimeError, match="No workbook path"):
        _calling_workbook_path()
